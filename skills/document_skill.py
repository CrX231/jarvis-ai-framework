import os
import tkinter as tk
from tkinter import filedialog

class DocumentSkill:
    def select_file(self):
        """Abre la ventana nativa de Windows para elegir un archivo."""
        # Inicializamos tkinter pero ocultamos su ventana principal fea
        root = tk.Tk()
        root.withdraw()
        # Forzamos que la ventana de selección aparezca por encima de todo
        root.attributes('-topmost', True) 
        
        # Abrimos el explorador de archivos
        ruta_archivo = filedialog.askopenfilename(
            title="Jarvis - Selecciona el documento a analizar",
            filetypes=[
                ("Todos los soportados", "*.pdf;*.docx;*.xlsx;*.txt;*.py;*.java;*.js;*.html;*.css;*.json;*.csv"),
                ("Documentos PDF", "*.pdf"),
                ("Word y Excel", "*.docx;*.xlsx"),
                ("Código y Texto", "*.txt;*.py;*.java;*.js;*.html;*.css;*.json;*.csv")
            ]
        )
        return ruta_archivo

    def analyze(self, command, brain, ruta=None):
        """Lee el archivo especificado o abre la ventana para elegirlo, extrae su texto y lo analiza."""
        if not ruta:
            ruta = self.select_file()
        
        if not ruta:
            return "Operación cancelada. No seleccionaste ningún documento."
            
        filename = os.path.basename(ruta)
        ext = os.path.splitext(filename)[1].lower()
        texto_extraido = ""
        
        try:
            # 1. TEXTO PLANO Y CÓDIGO
            if ext in ['.txt', '.py', '.java', '.js', '.html', '.css', '.json', '.csv']:
                with open(ruta, 'r', encoding='utf-8', errors='ignore') as f:
                    texto_extraido = f.read()
                    
            # 2. PDF
            elif ext == '.pdf':
                from pypdf import PdfReader
                reader = PdfReader(ruta)
                paginas = [page.extract_text() for page in reader.pages if page.extract_text()]
                texto_extraido = "\n".join(paginas)
                
            # 3. WORD
            elif ext == '.docx':
                import docx
                doc = docx.Document(ruta)
                texto_extraido = "\n".join([p.text for p in doc.paragraphs])
                
            # 4. EXCEL
            elif ext == '.xlsx':
                import openpyxl
                wb = openpyxl.load_workbook(ruta, data_only=True)
                lineas_excel = []
                for sheet in wb.sheetnames:
                    lineas_excel.append(f"--- Hoja: {sheet} ---")
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        if any(row):
                            lineas_excel.append(", ".join([str(c) if c is not None else "" for c in row]))
                texto_extraido = "\n".join(lineas_excel)
                
            else:
                return f"El formato {ext} no está soportado."
                
            if not texto_extraido.strip():
                return f"El archivo {filename} parece estar vacío o es una imagen sin texto seleccionable."
                
            # Protección contra archivos inmensos
            if len(texto_extraido) > 120000:
                texto_extraido = texto_extraido[:120000] + "\n... [Texto truncado] ..."

            super_prompt = (
                f"Analiza el contenido del archivo '{filename}':\n"
                f"========================================\n"
                f"{texto_extraido}\n"
                f"========================================\n"
                f"Teniendo ese archivo como contexto, responde a: {command}"
            )
            
            return brain.think(super_prompt)
            
        except Exception as e:
            return f"Hubo un error al procesar el documento: {e}"